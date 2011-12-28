#!/usr/bin/env python

import datetime

from csvkit.typeinference import NULL_TIME
from openpyxl.reader.excel import load_workbook

def sniff_dialect(path):
    return {}

def extract_column_names(path, dialect):
    book = load_workbook(path, use_iterators=True)
    sheet = book.get_active_sheet()
    headers = sheet.iter_rows().next()

    return headers

def normalize_date(dt):
    if dt.time() == NULL_TIME:
        return dt.date().isoformat()

    if dt.microsecond == 0:
        return dt.isoformat()

    ms = dt.microsecond

    if ms < 1000:
        return dt.replace(microsecond=0).isoformat()
    elif ms > 999000:
        return dt.replace(second=dt.second + 1, microsecond=0).isoformat()

    return dt.isoformat()

def sample_data(path, dialect, sample_size):
    book = load_workbook(path, use_iterators=True)
    sheet = book.get_active_sheet()

    samples = []

    for i, row in enumerate(sheet.iter_rows()):
        if i == 0:
            continue

        if i == sample_size + 1:
            break

        values = []

        for c in row:
            value = c.internal_value

            if value.__class__ is datetime.datetime:
                value = normalize_date(value)
            elif value.__class__ is float:
                if value % 1 == 0:
                    value = int(value)

            if value.__class__ in (datetime.datetime, datetime.date, datetime.time):
                value = value.isoformat()

            values.append(value)

        samples.append(values)

    return samples

