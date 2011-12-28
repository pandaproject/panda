#!/usr/bin/env python

import datetime
import logging
from math import floor

from django.conf import settings
from openpyxl.reader.excel import load_workbook

from redd import solr, utils
from redd.tasks.import_file import ImportFileTask

SOLR_ADD_BUFFER_SIZE = 500

class ImportXLSXTask(ImportFileTask):
    """
    Task to import all data for a dataset from an Excel/OpenOffice XLSX file.
    """
    name = 'redd.tasks.import.xlsx'

    def run(self, dataset_slug, external_id_field_index=None, *args, **kwargs):
        """
        Execute import.
        """
        from redd.models import Dataset
        
        log = logging.getLogger(self.name)
        log.info('Beginning import, dataset_slug: %s' % dataset_slug)

        dataset = Dataset.objects.get(slug=dataset_slug)

        task_status = dataset.current_task
        self.task_start(task_status, 'Preparing to import')

        book = load_workbook(dataset.data_upload.get_path(), use_iterators=True)
        sheet = book.get_active_sheet()
        row_count = sheet.get_highest_row()
        
        add_buffer = []

        for i, row in enumerate(sheet.iter_rows()):
            # Skip header
            if i == 0:
                continue

            values = []

            for c in row:
                value = c.internal_value

                if value.__class__ is datetime.datetime:
                    value = utils.xlsx.normalize_date(value)
                elif value.__class__ is float:
                    if value % 1 == 0:
                        value = int(value)

                if value.__class__ in (datetime.datetime, datetime.date, datetime.time):
                    value = value.isoformat()

                values.append(value)

            external_id = None

            if external_id_field_index is not None:
                external_id = values[external_id_field_index]

            data = utils.solr.make_data_row(dataset, values, external_id=external_id)

            add_buffer.append(data)

            if i % SOLR_ADD_BUFFER_SIZE == 0:
                solr.add(settings.SOLR_DATA_CORE, add_buffer)
                add_buffer = []

                task_status.message = '%.0f%% complete' % floor(float(i) / float(row_count) * 100)
                task_status.save()

                if self.is_aborted():
                    self.task_abort(self.task_status, 'Aborted after importing %.0f%%' % floor(float(i) / float(row_count) * 100))

                    log.warning('Import aborted, dataset_slug: %s' % dataset_slug)

                    return

        if add_buffer:
            solr.add(settings.SOLR_DATA_CORE, add_buffer)
            add_buffer = []

        solr.commit(settings.SOLR_DATA_CORE)

        self.task_update(task_status, '100% complete')

        dataset.row_count = row_count - 1
        dataset.save()

        log.info('Finished import, dataset_slug: %s' % dataset_slug)

