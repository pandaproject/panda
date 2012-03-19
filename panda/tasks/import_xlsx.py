#!/usr/bin/env python

import datetime
import logging
from math import floor

from django.conf import settings
from openpyxl.reader.excel import load_workbook

from panda import solr, utils
from panda.tasks.import_file import ImportFileTask
from panda.utils.typecoercion import DataTyper

SOLR_ADD_BUFFER_SIZE = 500

class ImportXLSXTask(ImportFileTask):
    """
    Task to import all data for a dataset from an Excel/OpenOffice XLSX file.
    """
    name = 'panda.tasks.import.xlsx'

    def run(self, dataset_slug, upload_id, external_id_field_index=None, *args, **kwargs):
        """
        Execute import.
        """
        from panda.models import Dataset, DataUpload
        
        log = logging.getLogger(self.name)
        log.info('Beginning import, dataset_slug: %s' % dataset_slug)

        dataset = Dataset.objects.get(slug=dataset_slug)
        upload = DataUpload.objects.get(id=upload_id)

        task_status = dataset.current_task
        task_status.begin('Preparing to import')

        book = load_workbook(upload.get_path(), use_iterators=True)
        sheet = book.get_active_sheet()
        row_count = sheet.get_highest_row()
        
        add_buffer = []
        data_typer = DataTyper(dataset.column_schema)

        for i, row in enumerate(sheet.iter_rows()):
            # Skip header
            if i == 0:
                continue

            values = []

            for c in row:
                value = c.internal_value

                if value.__class__ is datetime.datetime:
                    value = utils.xlsx.normalize_date(value)
                elif value.__class__ is float:
                    if value % 1 == 0:
                        value = int(value)

                if value.__class__ in (datetime.datetime, datetime.date, datetime.time):
                    value = value.isoformat()

                values.append(value)

            external_id = None

            if external_id_field_index is not None:
                external_id = values[external_id_field_index]

            data = utils.solr.make_data_row(dataset, values, external_id=external_id)
            data = data_typer(data, values)

            add_buffer.append(data)

            if i % SOLR_ADD_BUFFER_SIZE == 0:
                solr.add(settings.SOLR_DATA_CORE, add_buffer)
                add_buffer = []

                task_status.update('%.0f%% complete' % floor(float(i) / float(row_count) * 100))

                if self.is_aborted():
                    task_status.abort('Aborted after importing %.0f%%' % floor(float(i) / float(row_count) * 100))

                    log.warning('Import aborted, dataset_slug: %s' % dataset_slug)

                    return

        if add_buffer:
            solr.add(settings.SOLR_DATA_CORE, add_buffer)
            add_buffer = []

        solr.commit(settings.SOLR_DATA_CORE)

        task_status.update('100% complete')

        # Refresh dataset from database so there is no chance of crushing changes made since the task started
        dataset = Dataset.objects.get(slug=dataset_slug)

        if not dataset.row_count:
            dataset.row_count = i
        else:
            dataset.row_count += i
        
        dataset.column_schema = data_typer.schema

        dataset.save()

        # Refres
        upload = DataUpload.objects.get(id=upload_id)

        upload.imported = True
        upload.save()

        log.info('Finished import, dataset_slug: %s' % dataset_slug)

