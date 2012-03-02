#!/usr/bin/env python

import datetime
from itertools import islice
from types import NoneType

from csvkit.typeinference import NULL_TIME
from openpyxl.reader.excel import load_workbook

def sniff_dialect(path, **kwargs):
    return {}

def extract_column_names(path, dialect, **kwargs):
    book = load_workbook(path, use_iterators=True)
    sheet = book.get_active_sheet()
    headers = sheet.iter_rows().next()

    return [unicode(h.internal_value) if h.internal_value is not None else '' for h in headers]

def normalize_date(dt):
    if dt.time() == NULL_TIME:
        return dt.date().isoformat()

    if dt.microsecond == 0:
        return dt.isoformat()

    ms = dt.microsecond

    if ms < 1000:
        return dt.replace(microsecond=0).isoformat()
    elif ms > 999000:
        return dt.replace(second=dt.second + 1, microsecond=0).isoformat()

    return dt.isoformat()

def sample_data(path, dialect, sample_size, **kwargs):
    book = load_workbook(path, use_iterators=True)
    sheet = book.get_active_sheet()

    samples = []

    for i, row in enumerate(sheet.iter_rows()):
        if i == 0:
            continue

        if i == sample_size + 1:
            break

        values = []

        for c in row:
            value = c.internal_value

            if value.__class__ is datetime.datetime:
                value = normalize_date(value)
            elif value.__class__ is float:
                if value % 1 == 0:
                    value = int(value)

            if value.__class__ in (datetime.datetime, datetime.date, datetime.time):
                value = value.isoformat()

            values.append(unicode(value))

        samples.append(values)

    return samples

def determine_column_type(types):
    """
    Determine the correct type for a column from a list of cell types.
    """
    types_set = set(types)
    types_set.discard(NoneType)

    # Normalize mixed types to text
    if len(types_set) > 1:
        return unicode

    try:
        t = types_set.pop()
    except KeyError:
        return NoneType 

    # XLSX supports a time type, but Solr does not
    if t is datetime.time:
        return datetime.datetime
    else:
        return t 

def determine_number_type(values):
    """
    Determine if a column of numbers in an XLS file are integral.
    """
    # Test if all values are whole numbers, if so coerce floats it ints
    integral = True

    for v in values:
        if v and v % 1 != 0:
            integral = False
            break

    if integral:
        return int
    else:
        return float

def guess_column_types(path, dialect, sample_size, encoding='utf-8'):
    """
    Guess column types based on a sample of data.
    """
    book = load_workbook(path, use_iterators=True)
    sheet = book.get_active_sheet()

    rows = islice(sheet.iter_rows(), 0, sample_size + 1)
    rows.next()

    columns = zip(*rows)
    column_types = []

    for column in columns:
        values = [c.internal_value for c in column]

        t = determine_column_type([v.__class__ for v in values])

        if t is float:
            t = determine_number_type(values) 

        column_types.append(t)

    return [c.__name__ for c in column_types]

